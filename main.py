from bs4 import BeautifulSoup
import requests
import lxml
import openpyxl
from openpyxl import load_workbook

url_headers = {
    "Accept": "*/*",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
}


data_turt = []
monthlist = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']
citylist = ['norway/oslo', 'portugal/porto', 'montenegro/podgoroditsa', 'montenegro/bar',  'portugal/lisbon', 'canada/vancouver', 'argentina/buenos_aires', 'portugal/madeira', 'bulgaria/varna', 'russia/severodvinsk', 'new_zealand/auckland', 'serbia/belgrade', 'russia/moscow', 'czech_republic/prague']
citylist2 = ['Oslo','Porto','Podgorica','Bar-Montenegro','Lisbon',
        'Vancouver','Buenos-Aires','Funchal','Varna','Severodvinsk-Russia',
        'Auckland','Belgrade','Moscow','Prague']
pricelist = []

def collect_data_turt():
    count = 0
    for country_city in citylist:
        url_1 = f"https://pogoda.turtella.ru/{country_city}"

        for month in monthlist:
            url = f'{url_1}/{month}'
            response = requests.get(url=url, headers=url_headers)
            src = response.text
            soup = BeautifulSoup(src, "lxml")
            city = soup.find('div', class_="blockLnk bck big-btn").find("td").text
            month = soup.find('div', id="monthWeather").find('span').text
            avg_day = float(soup.find('div', id="monthWeather").find_all('tr')[1].find_all('td')[2].text[:-2])
            avg_night = float(soup.find('div', id="monthWeather").find_all('tr')[2].find_all('td')[1].text[:-2])
            try:
                sea_temp = float(soup.find('div', id="monthWeather").find_all('tr')[4].find_all('td')[2].text[:-2])
            except Exception:
                sea_temp = soup.find('div', id="monthWeather").find_all('tr')[4].find_all('td')[2].text[:-2]
            try:
                sun_days = float(soup.find('div', id="monthWeather").find_all('tr')[6].find_all('td')[2].text.split()[0])
            except Exception:
                sun_days = float(soup.find('div', id="monthWeather").find_all('tr')[4].find_all('td')[2].text.split()[0])
                sea_temp = 0
            data_turt.append(
                {
                    "Город": city,
                    "Месяц": month,
                    "Средняя температура днём": avg_day,
                    "Средняя температура ночью": avg_night,
                    "Кол-во солнечных дней": sun_days,
                    "Температура моря": sea_temp

                }
            )
        count +=1
        print(f'Обработан город № {count}')

    book = openpyxl.Workbook()  # создание файла
    sheet_1 = book.create_sheet("Данные")  # создание вкладки
    book.remove(book.active)  # удаление пустой вкладки

    headers = list(data_turt[0].keys()) #добавляем заголовки
    for col_num, header in enumerate(headers, 1):
        sheet_1.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(data_turt, 2):
        for col_num, cell_value in enumerate(row_data.values(), 1):
            sheet_1.cell(row=row_num, column=col_num, value=cell_value)
    book.save('data.xlsx')
    book.close()
    print("Файл записан, работа завершена")
def collect_data_numb():
    count = 0
    for city in citylist2:
        url = f'https://www.numbeo.com/cost-of-living/in/{city}?displayCurrency=USD'
        response = requests.get(url=url, headers=url_headers)
        src = response.text
        soup = BeautifulSoup(src, 'lxml')
        beef = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[16].find('span').text.split()[0])
        potato = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[21].find('span').text.split()[0])
        beer = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[26].find('span').text.split()[0])
        gasoline = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[35].find('span').text.split()[0])
        basic = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[39].find('span').text.split()[0])
        jeans = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[50].find('span').text.split()[0])
        rent = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[57].find('span').text.split()[
                0].replace(',', ''))
        price_psm = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[60].find('span').text.split()[
                0].replace(',', ''))
        salary = float(
            soup.find('table', class_="data_wide_table new_bar_table").find_all('tr')[63].find('span').text.split()[
                0].replace(',', ''))
        for i in range(12): #число нужно для того чтобы заполнять пустые строки при сохранении в файл
            pricelist.append(
                {
                    "Город": city,
                    "Говядина": beef,
                    "Картофель": potato,
                    "Пиво": beer,
                    "Бензин": gasoline,
                    "Джинсы": jeans,
                    "Аренда 3-к": rent,
                    "Коммунальные платежи": basic,
                    "Цена за квадратного метра": price_psm,
                    "Зарплата": salary

                }
            )
        count += 1
        print(f'Обработан город № {count}')

    book = load_workbook('data.xlsx')
    active_sheet = book.active #выбираем активную вкладку
    headers = list(pricelist[0].keys())  # добавляем заголовки
    for col_num, header in enumerate(headers, 7):# цифра семь это номер столбца с которого начинается запись
        active_sheet.cell(row=1, column=col_num, value=header)


        for row_num, row_data in enumerate(pricelist, 2):
            for col_num, cell_value in enumerate(row_data.values(), 7): # цифра семь это номер столбца с которого начинается запись
                active_sheet.cell(row=row_num, column=col_num, value=cell_value)

    book.save('data.xlsx')
    book.close()

    print("Файл записан, работа завершена")

def main():
    collect_data_turt()
    collect_data_numb()

if __name__=="__main__":
    main()
